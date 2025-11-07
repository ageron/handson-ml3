"""
Universal Import Registry v3.0
- Recursive scan (.py/.ipynb), excludes env folders
- Per-folder registries, central MASTER, DB snapshots with RunID
- Dashboard (Package usage, Outdated packages, Function/Class counts)
- Backups (keeps last 3)
"""

import ast
import nbformat
import importlib
import importlib.metadata
import pandas as pd
import requests
import sqlite3
import shutil
import getpass
import sys
import uuid
from packaging.version import parse as parse_version
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

# ---------------- CONFIG ----------------
# Workspace root to scan
# WORKSPACE_ROOT = Path(r"D:/aresources/workspace/mac-py/ml_ai")  # change if needed
WORKSPACE_ROOT = Path(r"D:/aresources/workspace/mac-py/mac-handson-ml3")  # changed location


# Central registry folder (you selected this earlier)
# CENTRAL_REGISTRY_DIR = Path(r"D:/aresources/workspace/registry_central")
CENTRAL_REGISTRY_DIR = Path(r"D:/aresources/workspace/mac-py/mac-handson-ml3")
CENTRAL_REGISTRY_DIR.mkdir(parents=True, exist_ok=True)

# DB and Excel names/paths
DB_PATH = CENTRAL_REGISTRY_DIR / "universal_registry.db"
MASTER_EXCEL = CENTRAL_REGISTRY_DIR / "universal_registry_MASTER.xlsx"
BACKUP_DIR = CENTRAL_REGISTRY_DIR / "backups"
BACKUP_DIR.mkdir(parents=True, exist_ok=True)
MAX_BACKUPS = 3
PYPI_TIMEOUT = 6

# Folders to skip during scanning
EXCLUDE_FOLDERS = {
    ".git", "__pycache__", ".ipynb_checkpoints", "venv", "env",
    ".conda", "site-packages", "build", "dist", ".pytest_cache"
}

# Columns expected in registry rows
REGISTRY_COLS = [
    "RunID","Timestamp","File","FilePath","FileType","Parent","Submodule",
    "Alias","Type","Category","Version","FunctionOrClass","Args",
    "Description","Lines","LastModified","SystemUser","PythonVersion"
]

PACKAGE_COLS = [
    "RunID","Timestamp","Parent","Version","UsageCount","LatestVersionPyPI",
    "CompareResult","Outdated","SystemUser","PythonVersion"
]

# ---------------- Helpers ----------------
def now_str(fmt="%Y-%m-%d %H:%M:%S"): return datetime.now().strftime(fmt)

def get_installed_version(pkg):
    try:
        return importlib.metadata.version(pkg.split(".")[0])
    except Exception:
        try:
            importlib.import_module(pkg.split(".")[0])
            return "builtin"
        except Exception:
            return "N/A"

def get_latest_pypi_version(package_name, timeout=PYPI_TIMEOUT):
    base = package_name.split('.')[0]
    url = f"https://pypi.org/pypi/{base}/json"
    try:
        r = requests.get(url, timeout=timeout)
        if r.status_code == 200:
            return r.json().get("info", {}).get("version")
    except Exception:
        pass
    return None

def compare_versions(installed, latest):
    if installed in (None, "N/A", "builtin") or not latest:
        return "unknown"
    try:
        iv = parse_version(installed)
        lv = parse_version(latest)
        if iv < lv: return "installed < latest"
        if iv == lv: return "installed == latest"
        return "installed > latest"
    except Exception:
        return "unknown"

def is_in_excluded_path(path: Path):
    return any(part in EXCLUDE_FOLDERS for part in path.parts)

# ---------------- Parsers ----------------
def parse_py_file(path: Path):
    rows = []
    try:
        src = path.read_text(encoding="utf-8", errors="ignore")
        tree = ast.parse(src)
    except Exception:
        return rows
    for node in ast.walk(tree):
        # Imports
        if isinstance(node, (ast.Import, ast.ImportFrom)):
            names = node.names if isinstance(node, ast.Import) else node.names
            module_base = None if isinstance(node, ast.Import) else (node.module or "")
            for alias in names:
                if isinstance(node, ast.Import):
                    parent = alias.name.split(".")[0]
                    sub = ".".join(alias.name.split(".")[1:]) or None
                else:
                    parent = (module_base.split(".")[0]) if module_base else None
                    sub = alias.name or None
                rows.append({
                    "Type":"Import",
                    "Parent": parent,
                    "Submodule": sub,
                    "Alias": alias.asname or None,
                    "Category":"module",
                    "FunctionOrClass": None,
                    "Args": None,
                    "Description": None,
                    "Lines": len(src.splitlines()),
                    "LastModified": datetime.fromtimestamp(path.stat().st_mtime) if path.exists() else None,
                    "File": path.name,
                    "FilePath": str(path.resolve()),
                    "FileType": "py"
                })
        # Functions
        elif isinstance(node, ast.FunctionDef):
            rows.append({
                "Type":"Function",
                "Parent": None,
                "Submodule": None,
                "Alias": None,
                "Category":"User Defined",
                "FunctionOrClass": node.name,
                "Args": ", ".join(a.arg for a in node.args.args),
                "Description": (ast.get_docstring(node) or "").split("\n")[0] if ast.get_docstring(node) else None,
                "Lines": len(src.splitlines()),
                "LastModified": datetime.fromtimestamp(path.stat().st_mtime) if path.exists() else None,
                "File": path.name,
                "FilePath": str(path.resolve()),
                "FileType":"py"
            })
        # Classes
        elif isinstance(node, ast.ClassDef):
            rows.append({
                "Type":"Class",
                "Parent": None,
                "Submodule": None,
                "Alias": None,
                "Category":"User Defined",
                "FunctionOrClass": node.name,
                "Args": None,
                "Description": (ast.get_docstring(node) or "").split("\n")[0] if ast.get_docstring(node) else None,
                "Lines": len(src.splitlines()),
                "LastModified": datetime.fromtimestamp(path.stat().st_mtime) if path.exists() else None,
                "File": path.name,
                "FilePath": str(path.resolve()),
                "FileType":"py"
            })
    return rows

def extract_code_from_notebook(nb_path: Path):
    try:
        with nb_path.open("r", encoding="utf-8") as f:
            nb = nbformat.read(f, as_version=4)
    except Exception:
        return []
    rows = []
    for cell in nb.cells:
        if cell.cell_type == "code" and cell.source.strip():
            # write cell source to temporary text for parsing
            tmp = nb_path.with_suffix(".temp.py")
            tmp.write_text(cell.source, encoding="utf-8")
            rows.extend(parse_py_file(tmp))
            tmp.unlink(missing_ok=True)
    return rows

# ---------------- Build per-folder registries ----------------
def build_registry_for_folder(folder: Path):
    if not folder.exists():
        return None
    local_reg_dir = folder / "universal_import_registry"
    local_reg_dir.mkdir(exist_ok=True)
    collected = []
    for f in folder.rglob("*"):
        if is_in_excluded_path(f): continue
        if f.suffix == ".py":
            collected.extend(parse_py_file(f))
        elif f.suffix == ".ipynb":
            collected.extend(extract_code_from_notebook(f))
    if not collected:
        print(f"No entries in {folder}")
        return None
    df = pd.DataFrame(collected)
    # normalize FilePath and dedupe
    df["FilePath"] = df["FilePath"].apply(lambda p: str(Path(p).resolve()))
    dedup_subset = ["FilePath","Type","Parent","Submodule","FunctionOrClass"]
    df = df.drop_duplicates(subset=[c for c in dedup_subset if c in df.columns], keep="first").reset_index(drop=True)
    # enrich version for imports
    df["Version"] = df["Parent"].apply(lambda p: get_installed_version(p) if pd.notna(p) else None)
    # add audit columns
    df["RunID"] = None
    df["Timestamp"] = None
    df["SystemUser"] = None
    df["PythonVersion"] = None
    out_file = local_reg_dir / f"universal_registry_{folder.name}.xlsx"
    df.to_excel(out_file, index=False)
    print(f"Saved local registry: {out_file}")
    return out_file

# ---------------- Merge locals into master and create dashboard ----------------
def merge_locals_and_create_master(root: Path):
    central = CENTRAL_REGISTRY_DIR
    local_files = list(root.rglob("universal_import_registry/universal_registry_*.xlsx"))
    if not local_files:
        print("No local registries found to merge.")
        return None
    dfs = []
    for lf in local_files:
        try:
            dfs.append(pd.read_excel(lf))
        except Exception as e:
            print(f"Could not read {lf}: {e}")
    if not dfs:
        return None
    master = pd.concat(dfs, ignore_index=True, sort=False)
    # normalize + dedupe
    if "FilePath" in master.columns:
        master["FilePath"] = master["FilePath"].apply(lambda p: str(Path(p).resolve()))
    dedup_cols = ["FilePath","Type","Parent","Submodule","FunctionOrClass"]
    master = master.drop_duplicates(subset=[c for c in dedup_cols if c in master.columns], keep="first").reset_index(drop=True)
    # Add run metadata
    run_id = str(uuid.uuid4())
    ts = now_str()
    master["RunID"] = run_id
    master["Timestamp"] = ts
    master["SystemUser"] = getpass.getuser()
    master["PythonVersion"] = sys.version.split()[0]
    # Package summary
    pkg_df = (master[master["Type"] == "Import"]
              .groupby(["Parent","Version"], dropna=False)
              .size().reset_index(name="UsageCount")
              .sort_values("UsageCount", ascending=False))
    pkg_df["LatestVersionPyPI"] = pkg_df["Parent"].apply(lambda p: get_latest_pypi_version(p) if pd.notna(p) else None)
    pkg_df["CompareResult"] = pkg_df.apply(lambda r: compare_versions(str(r["Version"]), r["LatestVersionPyPI"]), axis=1)
    pkg_df["Outdated"] = pkg_df["CompareResult"].apply(lambda x: "Yes" if x=="installed < latest" else ("No" if x and x.startswith("installed") else "Unknown"))
    # write master excel
    with pd.ExcelWriter(MASTER_EXCEL, engine="openpyxl") as writer:
        master.to_excel(writer, sheet_name="Registry", index=False)
        pkg_df.to_excel(writer, sheet_name="PackageSummary", index=False)
        # dashboard sheet
        summary = {
            "Metric":["Total Files Scanned","Unique Packages","Unique User Functions","Unique Classes","Registry Generated On","RunID","SystemUser"],
            "Value":[master["File"].nunique(), master.loc[master["Type"]=="Import","Parent"].nunique(),
                     master.loc[master["Type"]=="Function","FunctionOrClass"].nunique(),
                     master.loc[master["Type"]=="Class","FunctionOrClass"].nunique(), ts, run_id, getpass.getuser()]
        }
        pd.DataFrame(summary).to_excel(writer, sheet_name="Summary", index=False)
    print(f"Master Excel written: {MASTER_EXCEL}")
    return MASTER_EXCEL, run_id

# ---------------- Database (append-only, composite uniqueness) ----------------
def ensure_db(db_path: Path):
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    # RegistryHistory table with composite primary key (RunID, FilePath, Parent, Submodule, Type, FunctionOrClass)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS RegistryHistory (
        RunID TEXT,
        Timestamp TEXT,
        File TEXT,
        FilePath TEXT,
        FileType TEXT,
        Parent TEXT,
        Submodule TEXT,
        Alias TEXT,
        Type TEXT,
        Category TEXT,
        Version TEXT,
        FunctionOrClass TEXT,
        Args TEXT,
        Description TEXT,
        Lines INTEGER,
        LastModified TEXT,
        SystemUser TEXT,
        PythonVersion TEXT,
        PRIMARY KEY (RunID, FilePath, Parent, Submodule, Type, FunctionOrClass)
    )""")
    cur.execute("""
    CREATE TABLE IF NOT EXISTS PackageHistory (
        RunID TEXT,
        Timestamp TEXT,
        Parent TEXT,
        Version TEXT,
        UsageCount INTEGER,
        LatestVersionPyPI TEXT,
        CompareResult TEXT,
        Outdated TEXT,
        SystemUser TEXT,
        PythonVersion TEXT,
        PRIMARY KEY (RunID, Parent)
    )""")
    conn.commit()
    conn.close()

def backup_db(db_path: Path):
    if not db_path.exists():
        return
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = BACKUP_DIR / f"universal_registry_backup_{stamp}.db"
    shutil.copy2(db_path, dest)
    backups = sorted(BACKUP_DIR.glob("universal_registry_backup_*.db"), key=lambda p: p.stat().st_mtime, reverse=True)
    for old in backups[MAX_BACKUPS:]:
        try: old.unlink()
        except Exception: pass
    print(f"DB backup created: {dest}")

def insert_master_snapshot_to_db(master_excel: Path, run_id: str):
    ensure_db(DB_PATH)
    # backup current DB
    backup_db(DB_PATH)
    conn = sqlite3.connect(DB_PATH)
    xls = pd.ExcelFile(master_excel)
    df_master = pd.read_excel(xls, sheet_name="Registry")
    df_pkg = pd.read_excel(xls, sheet_name="PackageSummary")
    ts = now_str()
    # normalize columns
    df_master = df_master.fillna("")
    df_master["RunID"] = run_id
    df_master["Timestamp"] = ts
    df_master["SystemUser"] = getpass.getuser()
    df_master["PythonVersion"] = sys.version.split()[0]
    # prepare insert tuples for RegistryHistory
    reg_tuples = []
    for _, r in df_master.iterrows():
        reg_tuples.append((
            r.get("RunID"), r.get("Timestamp"), r.get("File"), r.get("FilePath"), r.get("FileType"),
            r.get("Parent"), r.get("Submodule"), r.get("Alias"), r.get("Type"), r.get("Category"),
            r.get("Version"), r.get("FunctionOrClass"), r.get("Args") if "Args" in r else None,
            r.get("Description"), int(r.get("Lines") or 0), str(r.get("LastModified") or ""), r.get("SystemUser"), r.get("PythonVersion")
        ))
    # insert with INSERT OR IGNORE (composite PK prevents duplicates)
    cur = conn.cursor()
    cur.executemany("""
        INSERT OR IGNORE INTO RegistryHistory
        (RunID,Timestamp,File,FilePath,FileType,Parent,Submodule,Alias,Type,Category,Version,FunctionOrClass,Args,Description,Lines,LastModified,SystemUser,PythonVersion)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, reg_tuples)

    # package tuples
    df_pkg = df_pkg.fillna("")
    df_pkg["RunID"] = run_id
    df_pkg["Timestamp"] = ts
    df_pkg["SystemUser"] = getpass.getuser()
    df_pkg["PythonVersion"] = sys.version.split()[0]
    pkg_tuples = []
    for _, r in df_pkg.iterrows():
        pkg_tuples.append((
            r.get("RunID"), r.get("Timestamp"), r.get("Parent"), r.get("Version"), int(r.get("UsageCount") or 0),
            r.get("LatestVersionPyPI"), r.get("CompareResult"), r.get("Outdated"), r.get("SystemUser"), r.get("PythonVersion")
        ))
    cur.executemany("""
        INSERT OR IGNORE INTO PackageHistory
        (RunID,Timestamp,Parent,Version,UsageCount,LatestVersionPyPI,CompareResult,Outdated,SystemUser,PythonVersion)
        VALUES (?,?,?,?,?,?,?,?,?,?)
    """, pkg_tuples)
    conn.commit()
    conn.close()
    print(f"DB snapshot inserted for RunID={run_id}")

# ---------------- Diff last two runs (console summary) ----------------
def diff_last_two_runs():
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("SELECT DISTINCT RunID, Timestamp FROM RegistryHistory ORDER BY Timestamp DESC LIMIT 2")
    runs = cur.fetchall()
    if len(runs) < 2:
        conn.close()
        print("Not enough runs for diff.")
        return
    new_run, old_run = runs[0][0], runs[1][0]
    df_new = pd.read_sql_query(f"SELECT Parent FROM RegistryHistory WHERE RunID='{new_run}'", conn)
    df_old = pd.read_sql_query(f"SELECT Parent FROM RegistryHistory WHERE RunID='{old_run}'", conn)
    conn.close()
    new = set(df_new["Parent"].dropna())
    old = set(df_old["Parent"].dropna())
    added = sorted(list(new - old))
    removed = sorted(list(old - new))
    print("New imports since last run:", added or "None")
    print("Removed imports since last run:", removed or "None")
    return {"added": added, "removed": removed}

# ---------------- Dashboard (adds a bar chart of top packages) ----------------
def add_dashboard_chart(master_excel_path: Path):
    try:
        wb = load_workbook(master_excel_path)
    except Exception:
        return
    if "PackageSummary" not in wb.sheetnames:
        return
    ws = wb["PackageSummary"]
    # ensure there's enough rows
    if ws.max_row < 2:
        wb.save(master_excel_path)
        return
    # create bar chart for top 10 packages by UsageCount
    data_ref = Reference(ws, min_col=3, min_row=1, max_row=min(11, ws.max_row))  # UsageCount is column 3 in our pkg_df write?
    # Unfortunately column indices depend on df layout; safer to find column numbers
    # We'll find UsageCount column index
    header = [cell.value for cell in ws[1]]
    if "UsageCount" in header:
        usage_col = header.index("UsageCount") + 1
        parent_col = header.index("Parent") + 1
        data = Reference(ws, min_col=usage_col, min_row=1, max_row=min(11, ws.max_row))
        cats = Reference(ws, min_col=parent_col, min_row=2, max_row=min(11, ws.max_row))
        chart = BarChart()
        chart.title = "Top packages by UsageCount"
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 20
        # put chart in a new sheet "DashboardChart"
        if "Dashboard" not in wb.sheetnames:
            wb.create_sheet("Dashboard")
        dash = wb["Dashboard"]
        dash.add_chart(chart, "B2")
        wb.save(master_excel_path)
        print("Dashboard chart added.")
    else:
        wb.save(master_excel_path)

# ---------------- Runner (full pipeline) ----------------
def run_pipeline(workspace_root: Path):
    # 1) build local registries for root and its immediate subfolders
    targets = [workspace_root] + [p for p in workspace_root.iterdir() if p.is_dir()]
    created = []
    for t in targets:
        print(f"Processing folder: {t}")
        try:
            out = build_registry_for_folder(t)
            if out: created.append(out)
        except Exception as e:
            print(f"Failed for {t}: {e}")

    # 2) merge all local registries into central MASTER
    master_excel, run_id = merge_locals_and_create_master(workspace_root)
    if not master_excel:
        print("No master created.")
        return

    # 3) insert to DB (append-only), with composite primary key to avoid duplicates
    insert_master_snapshot_to_db(master_excel, run_id)

    # 4) add dashboard chart to master excel
    add_dashboard_chart(master_excel)

    # 5) diff last two runs (console)
    diff_last_two_runs()

    print("Pipeline complete.")

# ----------------- Main -----------------
if __name__ == "__main__":
    run_pipeline(WORKSPACE_ROOT)
